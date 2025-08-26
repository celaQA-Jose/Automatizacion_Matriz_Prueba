import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import PieChart
from openpyxl.chart.reference import Reference
from io import BytesIO
from datetime import datetime
import os

LOGO_PATH = os.path.join(os.path.dirname(__file__), 'logo.png')

def generate_excel_report(data: dict) -> BytesIO:
    wb = openpyxl.Workbook()

    # Estilos comunes
    bold = Font(name='Arial', size=11, bold=True)
    bold10 = Font(name='Arial', size=10, bold=True)
    normal = Font(name='Arial', size=10)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(fill_type='solid', fgColor='DDDDDD')
    thin = Side(style='thin', color='000000')
    all_border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # HOJA 1: Instrucciones
    ws = wb.active
    ws.title = 'Instrucciones'
    ws.sheet_view.showGridLines = False

    # Cabecera A1
    ws['A1'] = 'Formato de matriz de prueba'
    ws['A1'].font = bold
    ws['A1'].alignment = center

    # Logo en B3:B5
    if os.path.exists(LOGO_PATH):
        img = Image(LOGO_PATH)
        img.width, img.height = 120, 120
        ws.merge_cells('B3:B5')
        ws.add_image(img, 'B3')
        for r in range(3, 6):
            cell = ws.cell(row=r, column=2)
            cell.border = border
            cell.alignment = center

    # Datos del proyecto
    ws.merge_cells('C3:D3')
    ws.cell(3, 3).value = 'Departamento de Sistemas'
    ws.cell(3, 3).font = bold
    ws.cell(3, 3).alignment = center
    for c in (3, 4): ws.cell(3, c).border = border

    ws.cell(3, 5).value = 'Código Wrike:'
    ws.cell(3, 5).font = bold; ws.cell(3, 5).alignment = right_align; ws.cell(3, 5).border = border
    ws.cell(3, 6).value = data.get('wrike', '')
    ws.cell(3, 6).font = normal; ws.cell(3, 6).alignment = left_align; ws.cell(3, 6).border = border

    ws.merge_cells('C4:D4')
    ws.cell(4, 3).value = 'Plan Pruebas Detallado'
    ws.cell(4, 3).font = bold; ws.cell(4, 3).alignment = center
    for c in (3, 4): ws.cell(4, c).border = border

    ws.cell(4, 5).value = 'Versión:'
    ws.cell(4, 5).font = bold; ws.cell(4, 5).alignment = right_align; ws.cell(4, 5).border = border
    ws.cell(4, 6).value = data.get('version', '')
    ws.cell(4, 6).font = normal; ws.cell(4, 6).alignment = left_align; ws.cell(4, 6).border = border

    ws.merge_cells('C5:D5')
    ws.cell(5, 3).value = f"{data.get('codigo','')} {data.get('nombre','')}"
    ws.cell(5, 3).font = bold; ws.cell(5, 3).alignment = center
    for c in (3, 4): ws.cell(5, c).border = border

    ws.cell(5, 5).value = 'Fecha:'
    ws.cell(5, 5).font = bold; ws.cell(5, 5).alignment = right_align; ws.cell(5, 5).border = border
    try:
        fecha_fmt = datetime.strptime(data.get('fecha_proyecto',''), '%Y-%m-%d').strftime('%d/%m/%Y')
    except:
        fecha_fmt = data.get('fecha_proyecto','')
    ws.cell(5, 6).value = fecha_fmt
    ws.cell(5, 6).font = normal; ws.cell(5, 6).alignment = left_align; ws.cell(5, 6).border = border

    ws.merge_cells('B6:F6')
    ws.cell(6, 2).value = 'Instrucciones / Recomendaciones'
    ws.cell(6, 2).font = bold; ws.cell(6, 2).alignment = center
    for c in range(2, 7): ws.cell(6, c).border = border

    # Ajuste columnas/filas Instrucciones
    for col, w in [('A',33.56),('B',33.78),('C',21.22),('D',39.78),('E',27.33),('F',21.78)]:
        ws.column_dimensions[col].width = w
    for r, h in {1:21, 2:21, 3:40.8, 4:24.6, 5:28.8, 6:24}.items():
        ws.row_dimensions[r].height = h

    # Etapas y descripciones largas
    ws['B13'] = 'Las 4 etapas de pruebas incluyen:'; ws['B13'].font = bold10; ws['B13'].alignment = center
    etapas = ['Planificacion:','Diseño:','Ejecucion:','Analisis de Resultados:']
    for i, text in enumerate(etapas, start=14):
        ws.cell(i, 3).value = text; ws.cell(i, 3).font = bold; ws.cell(i, 3).alignment = center
    textos_largos = {
        14: "Con base en las especificaciones de los cambios o nueva funcionalidad, se preparan los escenarios de prueba y todo lo relacionado a un plan de pruebas.  Los artefactos principales son"
        " a) Plan Master de Pruebas "
        "b) Plan Pruebas Detallado (lista de escenarios de prueba) "
        "c) Calendario o Actividades principales con fechas y asignaciones",
        15: "Se revisa la lista de escenarios del Plan de Pruebas Detallado y se elaboran los guiones (scripts) de prueba, indicando los resultados esperados.  Esta tarea se realiza con base en el análisis de requerimientos del proyecto o requerimiento.",
        16: "Se ejecutan las pruebas según calendario y pasos de los guiones de prueba.  Por cada sesión de prueba, se recomienda sacar una copia del guion original, y guardarlo con el nombre del escenario + la fecha de ejecución de la prueba.",
        17: "Se prepara el informe de pruebas y se obtienen las métricas de calidad por semana y por mes, para complementar dicho informe.  El informe lo genera el Test Manager y lo envía al Project Manager para complementar el informe de avance del proyecto."
    }
    for r, txt in textos_largos.items():
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        cell = ws.cell(r, 4)
        cell.value = txt
        cell.font = normal
        cell.alignment = left_align
        ws.row_dimensions[r].height = {14:135, 15:68, 16:80, 17:86}[r]

     # HOJA 2: Plan detallado
    #
    ws2 = wb.create_sheet('Plan detallado')
    ws2.sheet_view.showGridLines = False

    # Ajuste de columnas/filas en Plan detallado
    for col, w in {
        'A':9.11,'B':42,'C':92.56,'D':48,
        'E':27.67,'F':27.67,'G':22,'H':25,'I':25
    }.items():
        ws2.column_dimensions[col].width = w
    for r, h in {
        1:12.6, 2:36, 3:12.6, 4:12.6, 5:12.6,
        6:26.4, **{i:13 for i in range(7,16)}, 16:65
    }.items():
        ws2.row_dimensions[r].height = h

    # Logo en B2
    if os.path.exists(LOGO_PATH):
        img = Image(LOGO_PATH)
        img.width, img.height = 120, 120
        ws2.merge_cells('B2:B5')
        for row in ws2['B2:B5']:
            for c in row:
                c.border = all_border
        ws2.add_image(img, 'B1')
        for r in range(3, 6):
            cell = ws2.cell(row=r, column=2)
            cell.border = border
            cell.alignment = center

    # Encabezados fusionados (C2:D2, C3:D3, C4:D4)
    for row, text in [
        (2, 'Departamento de Sistemas'),
        (3, 'Plan Pruebas Detallado'),
        (4, 'Instrucciones / Recomendaciones')
    ]:
        ws2.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        cell = ws2.cell(row, 3)
        cell.value = text
        cell.font = bold
        cell.alignment = center
        for c in (3, 4):
            ws2.cell(row, c).border = border

    # Celda B5 con código+nombre
    cell = ws2.cell(row=5, column=3)
    ws2.merge_cells('C5:F5')
    for row in ws2['C5:F5']:
        for c in row:
            c.border = all_border
    cell.value = f"{data.get('codigo','')} {data.get('nombre','')}"
    cell.font = bold
    cell.alignment = center
    cell.border = border

    # Etiquetas en B6, B8–B13
    labels = {
        6:'Nombre del proyecto',
        8:'Codigo de Proyecto',
        9:'Sistema Impactado',
        10:'Modulo',
        11:'Fecha de elaboracion',
        12:'Fecha planead de ejecucion',
        13:'Elaborado por:'
    }
    for r, text in labels.items():
        c = ws2.cell(row=r, column=2)
        c.value = text
        c.font = bold10
        c.alignment = left_align

    ws2.cell(2, 5).value = 'Código Wrike:'
    ws2.cell(2, 5).font = bold; ws2.cell(2, 5).alignment = right_align; ws2.cell(2, 5).border = border
    ws2.cell(2, 6).value = data.get('wrike', '')
    ws2.cell(2, 6).font = normal; ws2.cell(2, 6).alignment = left_align; ws2.cell(2, 6).border = border

    ws2.cell(3, 5).value = 'Versión:'
    ws2.cell(3, 5).font = bold; ws2.cell(3, 5).alignment = right_align; ws2.cell(3, 5).border = border
    ws2.cell(3, 6).value = data.get('version', '')
    ws2.cell(3, 6).font = normal; ws2.cell(3, 6).alignment = left_align; ws2.cell(3, 6).border = border
 #Fecha Hoja 2, Plan Detallado
    ws2.cell(4, 5).value = 'Fecha:'
    ws2.cell(4, 5).font = bold; ws2.cell(4, 5).alignment = right_align; ws2.cell(4, 5).border = border
    try:
        fecha_fmt = datetime.strptime(data.get('fecha_proyecto',''), '%Y-%m-%d').strftime('%d/%m/%Y')
    except:
        fecha_fmt = data.get('fecha_proyecto','')
    ws2.cell(4, 6).value = fecha_fmt
    ws2.cell(4, 6).font = normal; ws2.cell(4, 6).alignment = left_align; ws2.cell(4, 6).border = border
    # ——— Aquí insertamos los valores en C6, C8–C12 ———
    # Nombre del proyecto en C6
    cell = ws2.cell(row=6, column=3)
    cell.value = data.get('nombre','')
    cell.font = normal
    cell.alignment = left_align

    # Código del proyecto en C8
    cell = ws2.cell(row=8, column=3)
    cell.value = data.get('codigo','')
    cell.font = normal
    cell.alignment = left_align

    # Sistema Impactado en C9
    cell = ws2.cell(row=9, column=3)
    cell.value = data.get('modulo','')
    cell.font = normal
    cell.alignment = left_align

    # Módulo en C10
    cell = ws2.cell(row=10, column=3)
    cell.value = data.get('modulo','')
    cell.font = normal
    cell.alignment = left_align

    # Fecha del Proyecto en C11
    try:
        f1 = datetime.strptime(data.get('fecha_proyecto',''), '%Y-%m-%d').strftime('%d/%m/%Y')
    except:
        f1 = data.get('fecha_proyecto','')
    cell = ws2.cell(row=11, column=3)
    cell.value = f1
    cell.font = normal
    cell.alignment = left_align

    # Fecha Planeada para ejecución en C12
    try:
        f2 = datetime.strptime(data.get('fecha_planeada',''), '%Y-%m-%d').strftime('%d/%m/%Y')
    except:
        f2 = data.get('fecha_planeada','')
    cell = ws2.cell(row=12, column=3)
    cell.value = f2
    cell.font = normal
    cell.alignment = left_align
    # ——————————————————————————————————————————————

    # Encabezado de la tabla de casos en fila 15
    start = 15
    headers = [
        'Código Prueba','Escenario','Descripción','Precondiciones',
        'Postcondiciones','Prioridad','Criterio de Aceptación','Comentarios'
    ]
    for idx, head in enumerate(headers, start=2):
        c = ws2.cell(row=start, column=idx)
        c.value = head
        c.font = bold
        c.alignment = center
        c.border = border
        c.fill = header_fill
    ws2.auto_filter.ref = f"B{start}:I{start}"
    ws2.freeze_panes = f"B{start+1}"

    # Datos de los casos desde la fila 16
    for case in data.get('casos', []):
        ws2.append([''] + [case[k] for k in ['codigo','caso','descripcion','precondiciones','postcondiciones','prioridad','criterio','comentarios']])
        for col in range(2, 10):
            c = ws2.cell(row=ws2.max_row, column=col)
            c.font = normal
            c.alignment = left_align
            c.border = border

        # Hojas individuales por caso
    for caso in data.get('casos', []):
        sheet_name = caso['codigo']
        ws_case = wb.create_sheet(title=sheet_name)
        ws_case.sheet_view.showGridLines = False

        # --- Header personalizado estilo Plan detallado ---
        # Logo en B2:B5
        if os.path.exists(LOGO_PATH):
            img = Image(LOGO_PATH); img.width=100; img.height=100
            # Anclar imagen a la celda B2 y ajustar tamaño a la celda
            img.anchor = 'B2'
            # Eliminar offsets para que se ajuste exactamente
            if hasattr(img.anchor, '_from'):
                img.anchor._from.colOff = 0
                img.anchor._from.rowOff = 0
            ws_case.merge_cells('B2:B6'); ws_case.add_image(img, 'B2')
            for r in range(2,6):
                ws_case.cell(row=r, column=2).border = border; ws_case.cell(row=r, column=2).alignment = center
        # Títulos C2:D2, C3:D3, C4:D4
        headers_meta = [(2,'Departamento de Sistemas'),(3,'Plan Pruebas Detallado'),(4,'Caso de Prueba')]
        for r, txt in headers_meta:
            ws_case.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            c = ws_case.cell(row=r, column=3)
            c.value = txt; c.font = bold; c.alignment = center
            for col in range(3,7): ws_case.cell(row=r, column=col).border = border
        # Código+Nombre en B6:F6
        ws_case.merge_cells('C5:F5')
        c = ws_case.cell(row=5, column=3)
        c.value = f"{caso['codigo']} {caso.get('caso','')}"; c.font=bold; c.alignment=center
        for col in range(2,7): ws_case.cell(row=5, column=col).border = border

        ws_case.merge_cells('C6:F6')
        c = ws_case.cell(row=6, column=3)
        for col in range(2,7): ws_case.cell(row=6, column=col).border = border

        ws_case.merge_cells('B7:F7')
        c = ws_case.cell(row=6, column=3)
        for col in range(2,7): ws_case.cell(row=6, column=col).border = border

        # Encabezados de tabla a partir de fila 8
        headers = ['Código Prueba','Escenario','Descripción','Precondiciones','Postcondiciones','Prioridad','Criterio de Aceptación','Comentarios']
        start_row = 8
        for idx, head in enumerate(headers, start=2):
            cell = ws_case.cell(row=start_row, column=idx)
            cell.value = head; cell.font = bold; cell.alignment = center; cell.border = border
        ws_case.freeze_panes = f'B{start_row+1}'

        # Datos fila 9
        keys = ['codigo','caso','descripcion','precondiciones','postcondiciones','prioridad','criterio','comentarios']
        for idx, key in enumerate(keys, start=2):
            cell = ws_case.cell(row=start_row+1, column=idx)
            cell.value = caso.get(key, ''); cell.font = normal; cell.alignment = left_align; cell.border = border

        # Ancho de columnas A-I
        for col in list('ABCDEFGHI'):
            ws_case.column_dimensions[col].width = 20

        # Sección de Evidencia de pruebas en B11
        # Fusionar B11:F11 si se desea todo el ancho
        ws_case.merge_cells('B11:F11')
        evid_cell = ws_case.cell(row=11, column=2)
        evid_cell.value = 'Evidencia de pruebas'
        evid_cell.font = bold
        evid_cell.alignment = center
        # Relleno azul claro
        evid_fill = PatternFill(fill_type='solid', fgColor='ADD8E6')  # light blue
        evid_cell.fill = evid_fill
        # Bordes en todos los lados
        for col in range(2, 7):
            cell = ws_case.cell(row=11, column=col)
            cell.border = border
            cell.fill = evid_fill

       # Agregar validación de lista en H9 para Criterio de Aceptación
        dv_sheet = DataValidation(
            type='list',
            formula1='"Aprobado,No aprobado,Pendiente,No Aplica"',
            allow_blank=True,
            showDropDown=True
        )
        ws_case.add_data_validation(dv_sheet)
        dv_sheet.add(f"H{start_row+1}")  # ahora en H9
        dv_sheet.add(cell.coordinate)


    # Nueva hoja: Resultados
    ws_result = wb.create_sheet(title='Resultados')
    ws_result.sheet_view.showGridLines = False

                # Preparar conteos dinámicos mediante fórmulas COUNTIF por hoja
    criterios = ['Pendiente', 'Aprobado', 'No Aprobado', 'No Aplica']
    ws_result.cell(1, 1, 'Criterio').font = bold
    ws_result.cell(1, 2, 'Cantidad').font = bold
    row = 2
    # Para cada criterio, construyo una fórmula que suma COUNTIF en cada hoja de caso
    for crit in criterios:
        ws_result.cell(row, 1, crit).font = normal
        # Generar partes de la fórmula para cada hoja
        parts = []
        for caso in data.get('casos', []):
            sheet_name = caso['codigo']
            # Cuenta el criterio en la columna I de cada hoja
            parts.append(f"COUNTIF('{sheet_name}'!H9, \"{crit}\" )")
        formula = '=' + '+'.join(parts) if parts else '0'
        cell = ws_result.cell(row, 2)
        cell.value = formula
        cell.font = normal
        cell.border = border
        cell.alignment = center
        row += 1
    # Total dinámico mediante SUMA de la columna Cantidad
    ws_result.cell(row, 1, 'Total').font = bold
    total_range = f"B2:B{row-1}"
    cell = ws_result.cell(row, 2)
    cell.value = f"=SUM({total_range})"
    cell.font = normal
    cell.border = border
    cell.alignment = center

    # Crear gráfico de pastel basado en datos (sin total)
    pie = PieChart()
    data_ref = Reference(ws_result, min_col=2, min_row=2, max_row=1 + len(criterios))
    labels = Reference(ws_result, min_col=1, min_row=2, max_row=1 + len(criterios))
    pie.add_data(data_ref, titles_from_data=False)
    pie.set_categories(labels)
    ws_result.add_chart(pie, 'D2')


    # Guardar y devolver buffer
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf